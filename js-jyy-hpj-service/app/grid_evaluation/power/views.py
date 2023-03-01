import json

from flask import request,Response
from . import comment
from app.utils.return_code import ReturnMessage, ErrorMessage, error_handler
from app.grid_evaluation.manager.generate_capacity_manager import GenerateCapacityManager
from app.grid_evaluation.manager.install_capacity_manager import InstallCapacityManager
from app.grid_evaluation.manager.social_power_and_load_value_manager import SocialPowerAndLoadValueManager
from app.grid_evaluation.manager.variable_capacity_manager import VariableCapacityManager
from app.grid_evaluation.manager.voltage_level_manager import VoltageLevelManager
from app.grid_evaluation.manager.per_house_capacity_manager import PerHouseCapacityManager

import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Font
from openpyxl.styles import Alignment
import string
import random
import os
import mimetypes
from werkzeug.datastructures import Headers
from urllib.parse import quote
from app.grid_evaluation.manager.excel_export_manager import ExcelExport
from tempfile import NamedTemporaryFile
from app.grid_evaluation.utils.indicator_utils import green_clean,flexible_intelligent
from app.grid_evaluation.utils.indicator_utils import type_indicator_renewable_power_rate
from app.grid_evaluation.utils.indicator_utils import type_indicator_water_power_rate
from app.grid_evaluation.utils.indicator_utils import type_indicator_wind_power_rate
from app.grid_evaluation.utils.indicator_utils import type_indicator_sun_power_rate
from app.grid_evaluation.utils.indicator_utils import type_indicator_biomass_power_rate
from app.grid_evaluation.utils.indicator_utils import type_indicator_renewable_install_rate
from app.grid_evaluation.utils.indicator_utils import type_indicator_water_install_rate
from app.grid_evaluation.utils.indicator_utils import type_indicator_wind_install_rate
from app.grid_evaluation.utils.indicator_utils import type_indicator_sun_install_rate
from app.grid_evaluation.utils.indicator_utils import type_indicator_biomass_install_rate
from app.grid_evaluation.utils.indicator_utils import type_indicator_reduce_co2
from app.grid_evaluation.utils.indicator_utils import type_indicator_reduce_so2
from app.grid_evaluation.utils.indicator_utils import type_indicator_wind_permeability
from app.grid_evaluation.utils.indicator_utils import type_indicator_sun_permeability
from app.grid_evaluation.utils.indicator_utils import type_indicator_social_power_year_grow_rate
from app.grid_evaluation.utils.indicator_utils import type_indicator_social_power_year_average_rate
from app.grid_evaluation.utils.indicator_utils import type_indicator_max_load_year_grow_rate
from app.grid_evaluation.utils.indicator_utils import type_indicator_max_load_year_average_rate
from app.grid_evaluation.utils.indicator_utils import type_indicator_variable_capacity_year_grow_rate
from app.grid_evaluation.utils.indicator_utils import type_indicator_variable_capacity_year_average_rate
from app.grid_evaluation.utils.indicator_utils import type_indicator_line_length_year_grow_rate
from app.grid_evaluation.utils.indicator_utils import type_indicator_line_length_year_average_rate
from app.grid_evaluation.utils.indicator_utils import type_indicator_unit_variable_capacity
from app.grid_evaluation.utils.indicator_utils import type_indicator_unit_line_length
from app.grid_evaluation.utils.indicator_utils import type_indicator_per_house_capacity
from app.grid_evaluation.utils.indicator_utils import type_indicator_hundred_average_single_line_length
from app.grid_evaluation.utils.indicator_utils import type_indicator_thirty_five_average_single_line_length
from app.grid_evaluation.utils.indicator_utils import type_indicator_net_connectivity
from app.grid_evaluation.utils.indicator_utils import type_indicator_gas_power_rate
from app.grid_evaluation.utils.indicator_utils import type_indicator_store_capacity_rate

from app.grid_evaluation.utils.indicator_utils import RENEWABLE_POWER_INSTALL, ENV_BENEFIT, NEW_POWER_PERMEABILITY
from app.grid_evaluation.utils.indicator_utils import NET_DEVELOP_SPEED, NET_SIZE, NET_FRAMEWORK, NET_ADJUST_ABILITY
from app.grid_evaluation.utils.indicator_utils import WATER_POWER, WIND_POWER, SUN_POWER, BIOMASS_POWER, SOCIAL_POWER
from app.grid_evaluation.utils.indicator_utils import RENEWABLE_POWER_RATE, WATER_POWER_RATE, WIND_POWER_RATE
from app.grid_evaluation.utils.indicator_utils import SUN_POWER_RATE, BIOMASS_POWER_RATE, TOTAL_POWER, WATER_INSTALL
from app.grid_evaluation.utils.indicator_utils import WIND_INSTALL, SUN_INSTALL, BIOMASS_INSTALL, TOTAL_INSTALL
from app.grid_evaluation.utils.indicator_utils import RENEWABLE_INSTALL_RATE, WATER_INSTALL_RATE, WIND_INSTALL_RATE
from app.grid_evaluation.utils.indicator_utils import SUN_INSTALL_RATE, BIOMASS_INSTALL_RATE, CO2_REDUCE, SO2_REDUCE
from app.grid_evaluation.utils.indicator_utils import WIND_PERMEABILITY, SUN_PERMEABILITY, SALE_POWER
from app.grid_evaluation.utils.indicator_utils import SALE_POWER_YEAR_GROW_RATE, SALE_POWER_YEAR_AVERAGE_RATE
from app.grid_evaluation.utils.indicator_utils import MAX_LOAD, MAX_LOAD_YEAR_GROW_RATE, MAX_LOAD_YEAR_AVERAGE_RATE
from app.grid_evaluation.utils.indicator_utils import VARIABLE_CAPACITY, VARIABLE_CAPACITY_YEAR_GROW_RATE
from app.grid_evaluation.utils.indicator_utils import VARIABLE_CAPACITY_YEAR_AVERAGE_RATE, LINE_LENGTH
from app.grid_evaluation.utils.indicator_utils import LINE_LENGTH_YEAR_GROW_RATE, LINE_LENGTH_YEAR_AVERAGE_RATE
from app.grid_evaluation.utils.indicator_utils import MAX_USE_LOAD, SUM_VARIABLE_CAPACITY, UNIT_VARIABLE_CAPACITY
from app.grid_evaluation.utils.indicator_utils import SUM_LINE_LENGTH, UNIT_LINE_LENGTH, PER_HOUSE_CAPACITY
from app.grid_evaluation.utils.indicator_utils import NET_CONNECTIVITY_LINE_NUMS, NET_CONNECTIVITY_SUBSTATION_NUMS
from app.grid_evaluation.utils.indicator_utils import NET_CONNECTIVITY, HUNDRED_AVERAGE_SINGLE_LINE_LENGTH
from app.grid_evaluation.utils.indicator_utils import HUNDRED_LINE_LENGTH_SUM, HUNDRED_LINE_NUMS
from app.grid_evaluation.utils.indicator_utils import THIRTY_FIVE_AVERAGE_SINGLE_LINE_LENGTH
from app.grid_evaluation.utils.indicator_utils import THIRTY_FIVE_LINE_LENGTH_SUM, THIRTY_FIVE_LINE_NUMS
from app.grid_evaluation.utils.indicator_utils import GAS_POWER_RATE, GAS_POWER, SALE_POWER
from app.grid_evaluation.utils.indicator_utils import STORE_CAPACITY_RATE, STORE_CAPACITY, NEW_ENERGY_INSTALL
from flask_login import login_required


@login_required
@comment.route('/menu', methods=['GET'])
@error_handler
def get_green_menu():
    try:
        name = request.args.get("name").strip()
        if name == '绿色清洁':
            return ReturnMessage.build_success(green_clean)
        if name == '灵活智能':
            return ReturnMessage.build_success(flexible_intelligent)
    except Exception as e:
        return ReturnMessage.build_error(ReturnMessage.INVALID_PARAM, e, 400)


@login_required
@comment.route('/power/names/search', methods=['GET'])
@error_handler
def get_power_search():
    try:
        indicator = request.args.get("indicator").strip()
        ret1 = []
        result = [
            {
                "type": RENEWABLE_POWER_INSTALL,
                "indicators": [{"indicator": RENEWABLE_POWER_RATE}]
            },
            {
                "type": RENEWABLE_POWER_INSTALL,
                "indicators": [{"indicator": WATER_POWER_RATE}],
            },
            {
                "type": RENEWABLE_POWER_INSTALL,
                "indicators": [{"indicator": WIND_POWER_RATE}],
            },
            {
                "type": RENEWABLE_POWER_INSTALL,
                "indicators": [{"indicator": SUN_POWER_RATE}],
            },
            {
                "type": RENEWABLE_POWER_INSTALL,
                "indicators": [{"indicator": BIOMASS_POWER_RATE}],
            },
            {
                "type": RENEWABLE_POWER_INSTALL,
                "indicators": [{"indicator": RENEWABLE_INSTALL_RATE}]},
            {
                "type": RENEWABLE_POWER_INSTALL,
                "indicators": [{"indicator": WATER_INSTALL_RATE}]},
            {
                "type": RENEWABLE_POWER_INSTALL,
                "indicators": [{"indicator": BIOMASS_INSTALL_RATE}]},
            {
                "type": RENEWABLE_POWER_INSTALL,
                "indicators": [{"indicator": WIND_INSTALL_RATE}]},
            {
                "type": RENEWABLE_POWER_INSTALL,
                "indicators": [{"indicator": SUN_INSTALL_RATE}]},
            {
                "type": ENV_BENEFIT,
                "indicators": [{"indicator": CO2_REDUCE}],
            },
            {
                "type": ENV_BENEFIT,
                "indicators": [{"indicator": CO2_REDUCE}],
            },
            {
                "type": NEW_POWER_PERMEABILITY,
                "indicators": [{"indicator": WIND_PERMEABILITY}],
            },
            {
                "type": NEW_POWER_PERMEABILITY,
                "indicators": [{"indicator": SUN_PERMEABILITY}],
            },
            {
                "type": NET_DEVELOP_SPEED,
                "indicators": [{"indicator": SALE_POWER_YEAR_GROW_RATE}]
            },
            {
                "type": NET_DEVELOP_SPEED,
                "indicators": [{"indicator": SALE_POWER_YEAR_AVERAGE_RATE}],
            },
            {
                "type": NET_DEVELOP_SPEED,
                "indicators": [{"indicator": MAX_LOAD_YEAR_GROW_RATE}],
            },
            {
                "type": NET_DEVELOP_SPEED,
                "indicators": [{"indicator": MAX_LOAD_YEAR_AVERAGE_RATE}],
            },
            {
                "type": NET_DEVELOP_SPEED,
                "indicators": [{"indicator": VARIABLE_CAPACITY_YEAR_GROW_RATE}],
            },
            {
                "type": NET_DEVELOP_SPEED,
                "indicators": [{"indicator": VARIABLE_CAPACITY_YEAR_AVERAGE_RATE}],
            },
            {
                "type": NET_DEVELOP_SPEED,
                "indicators": [{"indicator": LINE_LENGTH_YEAR_GROW_RATE}],
            },
            {
                "type": NET_DEVELOP_SPEED,
                "indicators": [{"indicator": LINE_LENGTH_YEAR_AVERAGE_RATE}],
            },
            {
                "type": NET_SIZE,
                "indicators": [{"indicator": UNIT_VARIABLE_CAPACITY}],
            },
            {
                "type": NET_SIZE,
                "indicators": [{"indicator": UNIT_LINE_LENGTH}],
            },
            {
                "type": NET_SIZE,
                "indicators": [{"indicator": PER_HOUSE_CAPACITY}],
            },
            {
                "type": NET_FRAMEWORK,
                "indicators": [{"indicator": HUNDRED_AVERAGE_SINGLE_LINE_LENGTH}],
            },
            {
                "type": NET_FRAMEWORK,
                "indicators": [{"indicator": THIRTY_FIVE_AVERAGE_SINGLE_LINE_LENGTH}]
            },
            {
                "type": NET_FRAMEWORK,
                "indicators": [{"indicator": NET_CONNECTIVITY}]},
            {
                "type": NET_ADJUST_ABILITY,
                "indicators": [{"indicator": GAS_POWER_RATE}]},
            {
                "type": NET_ADJUST_ABILITY,
                "indicators": [{"indicator": STORE_CAPACITY_RATE}],
            }
        ]
        for obj in result:
            if indicator in obj.get('indicators')[0].get('indicator'):
               ret1.append(obj)
        return ReturnMessage.build_success(ret1)

    except Exception as e:
        return ReturnMessage.build_error(ReturnMessage.INVALID_PARAM, e, 400)


@login_required
@comment.route('/power/names', methods=['GET'])
@error_handler
def get_power_names():
    try:
        type = request.args.get("type").strip()
        indicator = request.args.get("indicator").strip()
        if type == RENEWABLE_POWER_INSTALL and indicator == RENEWABLE_POWER_RATE:
            return ReturnMessage.build_success(type_indicator_renewable_power_rate)
        if type == RENEWABLE_POWER_INSTALL and indicator == WATER_POWER_RATE:
            return ReturnMessage.build_success(type_indicator_water_power_rate)
        if type == RENEWABLE_POWER_INSTALL and indicator == WIND_POWER_RATE:
            return ReturnMessage.build_success(type_indicator_wind_power_rate)
        if type == RENEWABLE_POWER_INSTALL and indicator == SUN_POWER_RATE:
            return ReturnMessage.build_success(type_indicator_sun_power_rate)
        if type == RENEWABLE_POWER_INSTALL and indicator == BIOMASS_POWER_RATE:
            return ReturnMessage.build_success(type_indicator_biomass_power_rate)
        if type == RENEWABLE_POWER_INSTALL and indicator == RENEWABLE_INSTALL_RATE:
            return ReturnMessage.build_success(type_indicator_renewable_install_rate)
        if type == RENEWABLE_POWER_INSTALL and indicator == WATER_INSTALL_RATE:
            return ReturnMessage.build_success(type_indicator_water_install_rate)
        if type == RENEWABLE_POWER_INSTALL and indicator == WIND_INSTALL_RATE:
            return ReturnMessage.build_success(type_indicator_wind_install_rate)
        if type == RENEWABLE_POWER_INSTALL and indicator == SUN_INSTALL_RATE:
            return ReturnMessage.build_success(type_indicator_sun_install_rate)
        if type == RENEWABLE_POWER_INSTALL and indicator == BIOMASS_INSTALL_RATE:
            return ReturnMessage.build_success(type_indicator_biomass_install_rate)
        if type == ENV_BENEFIT and indicator == CO2_REDUCE:
            return ReturnMessage.build_success(type_indicator_reduce_co2)
        if type == ENV_BENEFIT and indicator == SO2_REDUCE:
            return ReturnMessage.build_success(type_indicator_reduce_so2)
        if type == NEW_POWER_PERMEABILITY and indicator == WIND_PERMEABILITY:
            return ReturnMessage.build_success(type_indicator_wind_permeability)
        if type == NEW_POWER_PERMEABILITY and indicator == SUN_PERMEABILITY:
            return ReturnMessage.build_success(type_indicator_sun_permeability)
        if type == NET_DEVELOP_SPEED and indicator == SALE_POWER_YEAR_GROW_RATE:
            return ReturnMessage.build_success(type_indicator_social_power_year_grow_rate)
        if type == NET_DEVELOP_SPEED and indicator == SALE_POWER_YEAR_AVERAGE_RATE:
            return ReturnMessage.build_success(type_indicator_social_power_year_average_rate)
        if type == NET_DEVELOP_SPEED and indicator == MAX_LOAD_YEAR_GROW_RATE:
            return ReturnMessage.build_success(type_indicator_max_load_year_grow_rate)
        if type == NET_DEVELOP_SPEED and indicator == MAX_LOAD_YEAR_AVERAGE_RATE:
            return ReturnMessage.build_success(type_indicator_max_load_year_average_rate)
        if type == NET_DEVELOP_SPEED and indicator == VARIABLE_CAPACITY_YEAR_GROW_RATE:
            return ReturnMessage.build_success(type_indicator_variable_capacity_year_grow_rate)
        if type == NET_DEVELOP_SPEED and indicator == VARIABLE_CAPACITY_YEAR_AVERAGE_RATE:
            return ReturnMessage.build_success(type_indicator_variable_capacity_year_average_rate)
        if type == NET_DEVELOP_SPEED and indicator == LINE_LENGTH_YEAR_GROW_RATE:
            return ReturnMessage.build_success(type_indicator_line_length_year_grow_rate)
        if type == NET_DEVELOP_SPEED and indicator == LINE_LENGTH_YEAR_AVERAGE_RATE:
            return ReturnMessage.build_success(type_indicator_line_length_year_average_rate)
        if type == NET_SIZE and indicator == UNIT_VARIABLE_CAPACITY:
            return ReturnMessage.build_success(type_indicator_unit_variable_capacity)
        if type == NET_SIZE and indicator == UNIT_LINE_LENGTH:
            return ReturnMessage.build_success(type_indicator_unit_line_length)
        if type == NET_SIZE and indicator == PER_HOUSE_CAPACITY:
            return ReturnMessage.build_success(type_indicator_per_house_capacity)
        if type == NET_FRAMEWORK and indicator == HUNDRED_AVERAGE_SINGLE_LINE_LENGTH:
            return ReturnMessage.build_success(type_indicator_hundred_average_single_line_length)
        if type == NET_FRAMEWORK and indicator == THIRTY_FIVE_AVERAGE_SINGLE_LINE_LENGTH:
            return ReturnMessage.build_success(type_indicator_thirty_five_average_single_line_length)
        if type == NET_FRAMEWORK and indicator == NET_CONNECTIVITY:
            return ReturnMessage.build_success(type_indicator_net_connectivity)
        if type == NET_ADJUST_ABILITY and indicator == GAS_POWER_RATE:
            return ReturnMessage.build_success(type_indicator_gas_power_rate)
        if type == NET_ADJUST_ABILITY and indicator == STORE_CAPACITY_RATE:
            return ReturnMessage.build_success(type_indicator_store_capacity_rate)
    except Exception as e:
        return ReturnMessage.build_error(ReturnMessage.INVALID_PARAM, e, 400)


@login_required
@comment.route('/green/generate/detail', methods=['GET'])
@error_handler
def get_generate_power_detail():
    try:
        result = []
        name_list = request.args.get("name_list")
        if RENEWABLE_POWER_RATE in name_list:
            result.append(
                GenerateCapacityManager.get_generate_power_object_sum_detail()
            )
        if WATER_POWER_RATE in name_list:
            result.append(
                GenerateCapacityManager.get_generate_power_object_water_detail()
            )
        if WIND_POWER_RATE in name_list:
            result.append(
                GenerateCapacityManager.get_generate_power_object_wind_detail()
            )
        if SUN_POWER_RATE in name_list:
            result.append(
                GenerateCapacityManager.get_generate_power_object_sun_detail()
            )
        if BIOMASS_POWER_RATE in name_list:
            result.append(
                GenerateCapacityManager.get_generate_power_object_biomass_detail()
            )
        if WATER_POWER in name_list:
            query_raw = {'water_power': '水电发电量'}
            result.append(
                GenerateCapacityManager.get_generate_power_object_single_detail(query_raw)
            )
        if WIND_POWER in name_list:
            query_raw = {'wind_power': '风电发电量'}
            result.append(
                GenerateCapacityManager.get_generate_power_object_single_detail(query_raw)
            )
        if SUN_POWER in name_list:
            query_raw = {'sun_power': '太阳能发电量'}
            result.append(
                GenerateCapacityManager.get_generate_power_object_single_detail(query_raw)
            )
        if BIOMASS_POWER in name_list:
            query_raw = {'biomass_power': '生物质能发电量'}
            result.append(
                GenerateCapacityManager.get_generate_power_object_single_detail(query_raw)
            )
        if TOTAL_POWER in name_list:
            query_raw = {'total_power': '总发电量'}
            result.append(
                GenerateCapacityManager.get_generate_power_object_single_detail(query_raw)
            )
        if SOCIAL_POWER in name_list:
            query_raw = {'social_power': '全社会用电量'}
            result.append(
                GenerateCapacityManager.get_generate_power_object_single_detail(query_raw)
            )
        if WATER_INSTALL in name_list:
            query_raw = {'water_install': '水电装机容量'}
            result.append(
                InstallCapacityManager.get_install_capacity_object_single_detail(query_raw)
            )
        if WIND_INSTALL in name_list:
            query_raw = {'wind_install': '风电装机容量'}
            result.append(
                InstallCapacityManager.get_install_capacity_object_single_detail(query_raw)
            )
        if SUN_INSTALL in name_list:
            query_raw = {'sun_install': '太阳能装机容量'}
            result.append(
                InstallCapacityManager.get_install_capacity_object_single_detail(query_raw)
            )
        if BIOMASS_INSTALL in name_list:
            query_raw = {'biomass_install': '生物质能装机容量'}
            result.append(
                InstallCapacityManager.get_install_capacity_object_single_detail(query_raw)
            )
        if TOTAL_INSTALL in name_list:
            query_raw = {'total_install': '并网发电装机容量'}
            result.append(
                InstallCapacityManager.get_install_capacity_object_single_detail(query_raw)
            )
        if RENEWABLE_INSTALL_RATE in name_list:
            result.append(
                InstallCapacityManager.get_install_capacity_object_sum_detail()
            )
        if WATER_INSTALL_RATE in name_list:
            result.append(
                InstallCapacityManager.get_install_capacity_object_water_detail()
            )
        if WIND_INSTALL_RATE in name_list:
            result.append(
                InstallCapacityManager.get_install_capacity_object_wind_detail()
            )
        if SUN_INSTALL_RATE in name_list:
            result.append(
                InstallCapacityManager.get_install_capacity_object_sun_detail()
            )
        if BIOMASS_INSTALL_RATE in name_list:
            result.append(
                InstallCapacityManager.get_install_capacity_object_biomass_detail()
            )
        if CO2_REDUCE in name_list:
            result.append(
                GenerateCapacityManager.get_reduce_co2_detail()
            )
        if SO2_REDUCE in name_list:
            result.append(
                GenerateCapacityManager.get_reduce_so2_detail()
            )
        if WIND_PERMEABILITY in name_list:
            result.append(
                GenerateCapacityManager.get_wind_permeability_detail()
            )
        if SUN_PERMEABILITY in name_list:
            result.append(
                GenerateCapacityManager.get_sun_permeability_detail()
            )
        if SALE_POWER in name_list:
            result.append(
                SocialPowerAndLoadValueManager.get_sale_power_single_from_name()
            )
        if SALE_POWER_YEAR_GROW_RATE in name_list:
            result.append(
                SocialPowerAndLoadValueManager.get_sale_power_year_grow_rate_from_name()
            )
        if SALE_POWER_YEAR_AVERAGE_RATE in name_list:
            result.append(
                SocialPowerAndLoadValueManager.get_sale_power_average_grow_rate_from_name()
            )
        if MAX_LOAD in name_list:
            result.append(
                SocialPowerAndLoadValueManager.get_load_value_single_from_name()
            )
        if MAX_LOAD_YEAR_GROW_RATE in name_list:
            result.append(
                SocialPowerAndLoadValueManager.get_load_value_year_grow_rate_from_name()
            )
        if MAX_LOAD_YEAR_AVERAGE_RATE in name_list:
            result.append(
                SocialPowerAndLoadValueManager.get_load_value_average_grow_rate_from_name()
            )
        if VARIABLE_CAPACITY in name_list:
            result.append(
                VariableCapacityManager.get_variable_capacity_single_from_name()
            )
        if VARIABLE_CAPACITY_YEAR_GROW_RATE in name_list:
            result.append(
                VariableCapacityManager.get_variable_capacity_year_grow_from_name()
            )
        if VARIABLE_CAPACITY_YEAR_AVERAGE_RATE in name_list:
            result.append(
                VariableCapacityManager.get_variable_capacity_average_year_grow_from_name()
            )
        if LINE_LENGTH in name_list:
            result.append(
                VoltageLevelManager.get_length_single_from_name()
            )
        if LINE_LENGTH_YEAR_GROW_RATE in name_list:
            result.append(
                VoltageLevelManager.get_length_year_grow_from_name()
            )
        if LINE_LENGTH_YEAR_AVERAGE_RATE in name_list:
            result.append(
                VoltageLevelManager.get_length_average_year_grow_from_name()
            )
        if MAX_USE_LOAD in name_list:
            result.append(
                SocialPowerAndLoadValueManager.unit_high_load_single()
            )
        if SUM_VARIABLE_CAPACITY in name_list:
            result.append(
                SocialPowerAndLoadValueManager.unit_variable_capacity_single()
            )
        if SUM_LINE_LENGTH in name_list:
            result.append(
                SocialPowerAndLoadValueManager.unit_length_load_length_sum()
            )
        if PER_HOUSE_CAPACITY in name_list:
            result.append(
                PerHouseCapacityManager.get_average_house()
            )
        if HUNDRED_LINE_LENGTH_SUM in name_list:
            result.append(
                VoltageLevelManager.get_110kv_length()
            )
        if HUNDRED_LINE_NUMS in name_list:
            result.append(
                VoltageLevelManager.get_110kv_nums()
            )
        if THIRTY_FIVE_LINE_LENGTH_SUM in name_list:
            result.append(
                VoltageLevelManager.get_35kv_length()
            )
        if THIRTY_FIVE_LINE_NUMS in name_list:
            result.append(
                VoltageLevelManager.get_35kv_nums()
            )
        if NET_CONNECTIVITY_LINE_NUMS in name_list:
            result.append(
                VoltageLevelManager.get_connectivity_line_nums()
            )
        if NET_CONNECTIVITY_SUBSTATION_NUMS in name_list:
            result.append(
                VoltageLevelManager.get_connectivity_substation_nums()
            )
        if NET_CONNECTIVITY in name_list:
            result.append(
                VoltageLevelManager.get_connectivity()
            )
        if GAS_POWER in name_list:
            result.append(
                SocialPowerAndLoadValueManager.get_flexible_gas_power_single_detail()
            )
        if STORE_CAPACITY in name_list:
            result.append(
                SocialPowerAndLoadValueManager.get_store_capacity_single_detail()
            )
        if NEW_ENERGY_INSTALL in name_list:
            result.append(
                SocialPowerAndLoadValueManager.get_new_energy_capacity_single_detail()
            )
        if UNIT_VARIABLE_CAPACITY in name_list:
            result.append(
                SocialPowerAndLoadValueManager.unit_variable_load()
            )
        if UNIT_LINE_LENGTH in name_list:
            result.append(
                SocialPowerAndLoadValueManager.unit_length_load()
            )
        if HUNDRED_AVERAGE_SINGLE_LINE_LENGTH in name_list:
            result.append(
                VoltageLevelManager.get_110kv_average_single_length()
            )
        if THIRTY_FIVE_AVERAGE_SINGLE_LINE_LENGTH in name_list:
            result.append(
                VoltageLevelManager.get_35kv_average_single_length()
            )
        if GAS_POWER_RATE in name_list:
            result.append(
                SocialPowerAndLoadValueManager.get_flexible_gas_power_detail()
            )
        if STORE_CAPACITY_RATE in name_list:
            result.append(
                SocialPowerAndLoadValueManager.get_new_energy_capacity_rate_detail()
            )

    except Exception as e:
        return ReturnMessage.build_error(ReturnMessage.INVALID_PARAM, e, 400)

    return ReturnMessage.build_success(result)


def http_download(stream, filename):
    response = Response()
    response.status_code = 200
    response.data = stream
    mimetypes_tuple = mimetypes.guess_type(filename)
    response_headers = Headers({
        'Pragma': 'public',
        'Expires': '0',
        'Cache-Control': 'must-revalidate,post-check=0,pre-check=0',
        'Content-Type': 'application/octet-stream',
        'Content-Disposition': 'attachment;filename=\"%s\";' % quote(filename),
        'Content-Transfer-Encoding': 'binary',
        'Content-Length': len(response.data)

    })
    response.headers = response_headers
    return response


@login_required
@comment.route('/download', methods=['POST'])
def download():
    tmp_file = ""
    file_name = ""
    try:
        raw_data = request.get_data()
        data2 = json.loads(raw_data)
        file_data = data2.get('base64String')
        data = data2.get('resultList')

        wb = openpyxl.Workbook()
        ws2 = wb.create_sheet('Data')

        ws2.row_dimensions[1].height = 100
        a = 1
        ws2.cell(1, 1, '时间')
        ws2.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
        ws2.cell(row=1, column=1).font = Font(name='微软雅黑', size=10, bold=True, italic=False, underline='none',
                                              color='000000')
        for d in data:
            ws2.cell(1, a + 1, d.get("indicator"))
            ws2.cell(row=1, column=a + 1).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
            ws2.cell(row=1, column=a + 1).font = Font(name='微软雅黑', size=10, bold=True, italic=False, underline='none',
                                                      color='000000')
            a += 1

        b2 = 2
        list1 = data[0].get('resultList')
        for l in list1:
            ws2.cell(b2, 1, l.get('time'))
            b2 += 1

        cols2 = len(data) + 2
        a3 = 2
        for d in data:
            while a3 < cols2:
                list2 = d.get('resultList')
                b3 = 2
                for l in list2:
                    ws2.cell(b3, a3, l.get('value'))
                    b3 += 1
                break
            a3 += 1
        rows2 = len(data[0].get('resultList'))

        list11 = [rows2 + 2, 1, rows2 + 5, cols2]

        ws2.merge_cells(start_row=list11[0], start_column=list11[1], end_row=list11[2], end_column=list11[3] - 1)

        random_str = ''.join(random.sample(string.digits * 5 + string.ascii_letters * 4, 20))
        file_name = 'app/grid_evaluation/power/' + random_str + '.png'
        ExcelExport.base64_to_photo(file_data, file_name)
        img = Image(file_name)
        a = list11[3] - list11[1]
        img.width, img.height = 68 * a, 75
        ws2.add_image(img, 'A{}'.format(list11[0]))

        with NamedTemporaryFile(delete=False) as tmp:
            tmp_file = tmp.name
            wb.save(tmp.name)
            tmp.seek(0)
            output = tmp.read()

        return http_download(output, '未命名.xlsx')
    except Exception as e:
        return ReturnMessage.build_error(ReturnMessage.INVALID_PARAM, e, 400)
    finally:
        if os.path.exists(tmp_file):
            os.remove(tmp_file)
            print('成功删除文件:', tmp_file)
        else:
            print('未找到此文件:', tmp_file)
        if os.path.exists(file_name):
            os.remove(file_name)
            print('成功删除文件:', file_name)
        else:
            print('未找到此文件:', file_name)
