import openpyxl
from app.project_evaluation.manager import ProjectDetailManager, ProjectComponentManager, ProjectManager
from io import BytesIO


class DataParser:

    def check_new_project_format(self, file):
        wb = openpyxl.load_workbook(file, data_only=True)
        if "1 项目基本信息" not in wb or \
                "1.1项目建设过程信息" not in wb or \
                "1.2项目招标、合同管理与执行信息表" not in wb or \
                "1.3项目投资控制信息表" not in wb or \
                "1.4项目运行安全信息" not in wb or \
                "4、输变电工程决算基础数据统计" not in wb or \
                "5、输变电工程运营年度电量统计与预测" not in wb or \
                "6、工程历年运营期成本费用" not in wb or \
                "7、工程经济效益指标一览表" not in wb or \
                "2 项目投产主变信息" not in wb or \
                "3 项目投产线路信息" not in wb:
            raise Exception("Template Wrong!")
        for project in self.parse_sheet(wb["1 项目基本信息"]):
            if not project.get("WBS编码"):
                continue
            if "项目名称" not in project or \
                    "电压等级" not in project or \
                    "实际投运年" not in project or \
                    "工程分类" not in project:
                raise Exception("Project require field  Wrong!")

    def parse_new_project(self, file):
        project_wbscode = set()

        wb = openpyxl.load_workbook(file, data_only=True)
        for project in self.parse_sheet(wb["1 项目基本信息"]):
            wbs_code = project.get("WBS编码")
            if wbs_code:
                wbs_code = str(wbs_code).strip()
                project_wbscode.add(wbs_code)
                ProjectManager.add_project(
                    name=project["项目名称"],
                    wbs_code=wbs_code,
                    voltage_level=project["电压等级"],
                    operation_year=project["实际投运年"],
                    classification=project["工程分类"],
                    plan_year=project.get("规划投运年"),
                    company_province=project.get("所属省公司名称"),
                    company_city=project.get("所属地市公司名称"),
                    company_county=project.get("所属县公司名称"),
                    energy_type=project.get("“保障电源送出”和“服务新能源”项目电源类型"),
                    channel_type=project.get("“加强输电通道类”项目的输电通道类型"),
                    budget=project.get("决算投资（含税）（万元）"),
                    capacitance=project.get("新增变电容量（MVA）"),
                    wire_length=project.get("新增线路长度（公里）"),
                    transmission_capacity=project.get("新增输电能力（MW）")
                )

        for sheet_name in ["1.1项目建设过程信息", "1.2项目招标、合同管理与执行信息表", "1.3项目投资控制信息表", "1.4项目运行安全信息",
                           "4、输变电工程决算基础数据统计", "5、输变电工程运营年度电量统计与预测", "6、工程历年运营期成本费用",
                           "7、工程经济效益指标一览表"]:
            for project_detail in self.parse_sheet(wb[sheet_name]):
                wbs_code = project_detail.get("WBS编码")
                if wbs_code:
                    wbs_code = str(wbs_code).strip()
                    project_wbscode.add(wbs_code)
                    del project_detail["WBS编码"]
                    if sheet_name == "1.4项目运行安全信息":
                        self.aggregate_param(project_detail, "母线电压合格率（%）")
                        self.aggregate_param(project_detail, "继电保护和安稳装置误动、拒动次数（次）")
                        self.aggregate_param(project_detail, "影响电能质量考核次数（次）")
                    if sheet_name == "5、输变电工程运营年度电量统计与预测":
                        self.aggregate_param(project_detail, "工程供（输）电量/工程供（输）电量（万kWh）")
                        self.aggregate_param(project_detail, "工程供（输）电量/方案2预测值（万kWh）")
                    if sheet_name == "6、工程历年运营期成本费用":
                        self.aggregate_param(project_detail, "经营成本（总计）（万元）")

                    ProjectDetailManager.add_project_detail(wbs_code, project_detail)

        for component in self.parse_sheet(wb["2 项目投产主变信息"]):
            wbs_code = component.get("WBS编码")
            name = component.get("变电站名称", "") + component.get("主变名称", "")
            if wbs_code and name:
                wbs_code = str(wbs_code).strip()
                project_wbscode.add(wbs_code)
                del component["WBS编码"], component["主变名称"], component["变电站名称"]
                for aggregate_field in ["变电站最大负荷时刻有功功率（MW）", "上网电量（万kWh）", "上网电量（万kWh）",
                                        "变电站最大负荷时刻无功功率（Mvar）", "变电站最大负荷时刻无功功率（Mvar）",
                                        "变电损耗电量（万kWh）", "变压器强迫停运次数（次）", "变压器强迫停运时间（小时）"]:
                    self.aggregate_param(component, aggregate_field)

                ProjectComponentManager.add_project_component(wbs_code=wbs_code, name=name, params=component)

        for component in self.parse_sheet(wb["3 项目投产线路信息"]):
            wbs_code = component.get("WBS编码")
            name = component.get("线路名称")
            if name and wbs_code:
                wbs_code = str(wbs_code).strip()
                project_wbscode.add(wbs_code)
                del component["WBS编码"], component["线路名称"]
                for aggregate_field in ["线路最大负荷时刻有功功率（MW）", "线路最大负荷时刻无功功率（Mvar）",
                                        "正向输送电量（万kWh）", "反向输送电量（万kWh）", "线路损耗电量（万kWh）",
                                        "线路强迫停运次数（次）", "线路强迫停运时间（小时）"]:
                    self.aggregate_param(component, aggregate_field)

                ProjectComponentManager.add_project_component(wbs_code=str(wbs_code), name=name, params=component)

        return project_wbscode

    def check_old_project_format(self, file):
        wb = openpyxl.load_workbook(file, data_only=True)
        if "1 项目基本信息" not in wb or \
                "2 项目投产主变信息" not in wb or \
                "3 项目投产线路信息" not in wb:
            raise Exception("Template Wrong!")
        for project in self.parse_sheet(wb["1 项目基本信息"]):
            if not project.get("1 项目基本信息/WBS编码"):
                continue
            if "1 项目基本信息/项目名称" not in project or \
                    "1 项目基本信息/电压等级" not in project or \
                    "1 项目基本信息/实际投运年" not in project or \
                    "1 项目基本信息/工程分类" not in project:
                raise Exception("Project require field  Wrong!")

    def parse_old_project(self, file):
        wb = openpyxl.load_workbook(file, data_only=True)
        header = self.get_header(wb["1 项目基本信息"])
        year_list = []
        for name in header[0]:
            if name.startswith("5、输变电工程运营年度电量统计与预测/方案2预测值（万kWh）/"):
                year_list.append(name.split("/")[-1])
        year = str(int(sorted(year_list)[0]) - 1)

        project_wbscode = set()
        for project in self.parse_sheet(wb["1 项目基本信息"]):
            wbs_code = project.get("1 项目基本信息/WBS编码")
            if wbs_code:
                wbs_code = str(wbs_code).strip()
                project_wbscode.add(wbs_code)
                ProjectManager.add_project(
                    name=project["1 项目基本信息/项目名称"],
                    wbs_code=wbs_code,
                    voltage_level=project["1 项目基本信息/电压等级"],
                    operation_year=project["1 项目基本信息/实际投运年"],
                    classification=project["1 项目基本信息/工程分类"],
                    plan_year=project.get("1 项目基本信息/规划投运年"),
                    company_province=project.get("1 项目基本信息/所属省公司名称"),
                    company_city=project.get("1 项目基本信息/所属地市公司名称"),
                    company_county=project.get("1 项目基本信息/所属县公司名称"),
                    channel_type=project.get("1 项目基本信息/“加强输电通道类”项目的输电通道类型"),
                    energy_type=None,
                    budget=None,
                    capacitance=None,
                    wire_length=None,
                    transmission_capacity=None
                )

                project_detail = {}
                for name, value in project.items():
                    name = "/".join(name.split("/")[1:])
                    project_detail[name] = value
                self.aggregate_param(project_detail, "方案2预测值（万kWh）")
                for name, value in project_detail.items():
                    if name in ("母线电压合格率（%）", "继电保护和安稳装置误动、拒动次数（次）", "影响电能质量考核次数（次）",
                                "累计并网装机容量（MW）", "上网电量（万kWh）", "下网电量（万kWh）"):
                        ProjectDetailManager.add_project_detail(wbs_code, {name: [{"year": year, "value": value}]})
                    if name in ("工程供（输）电量（万kWh）", "方案3预测值（万kWh）"):
                        ProjectDetailManager.add_project_detail(wbs_code, {"工程供（输）电量/工程供（输）电量（万kWh）":
                                                                               [{"year": year, "value": value}]})
                    if name == "方案2预测值（万kWh）":
                        ProjectDetailManager.add_project_detail(wbs_code, {"工程供（输）电量/方案2预测值（万kWh）": value})
                    if name in ("方案二/总投资内部收益率（税后）（%）", "方案二/资本金内部收益率（税后）（%）", "方案二/静态投资回收期（年）",
                                "方案二/偿债备付率", "方案二/利息备付率"):
                        ProjectDetailManager.add_project_detail(wbs_code, {name: value})

        for component in self.parse_sheet(wb["2 项目投产主变信息"]):
            wbs_code = component.get("WBS编码")
            component_name = component.get("变电站名称", "") + component.get("主变名称", "")
            if wbs_code and component_name:
                wbs_code = str(wbs_code).strip()
                project_wbscode.add(wbs_code)
                project_component = {}
                for name, value in component.items():
                    name = "/".join(name.split("/")[1:])
                    project_component[name] = value

                for name, value in project_component.items():
                    if name in ("变电站最大负荷时刻有功功率（MW）", "变电站最大负荷时刻无功功率（Mvar）", "电网最大负荷时刻有功功率（MW）",
                                "电网最大负荷时刻无功功率（Mvar）", "上网电量（万kWh）", "下网电量（万kWh）", "变电损耗电量（万kWh）",
                                "变压器强迫停运次数（次）", "变压器强迫停运时间（小时）"):
                        ProjectComponentManager.add_project_component(wbs_code=wbs_code,
                                                                      name=component_name,
                                                                      params={name: [{"year": year, "value": value}]})

        for component in self.parse_sheet(wb["3 项目投产线路信息"]):
            wbs_code = component.get("WBS编码")
            component_name = component.get("线路名称")
            if wbs_code and component_name:
                wbs_code = str(wbs_code).strip()
                project_wbscode.add(wbs_code)
                project_component = {}
                for name, value in component.items():
                    name = "/".join(name.split("/")[1:])
                    project_component[name] = value

                for name, value in project_component.items():
                    if name in ("线路最大负荷时刻有功功率（MW）", "线路最大负荷时刻无功功率（Mvar）", "电网最大负荷时刻有功功率（MW）",
                                "电网最大负荷时刻无功功率（Mvar）", "正向输送电量（万kWh）", "反向输送电量（万kWh）", "线路损耗电量（万kWh）",
                                "线路强迫停运次数（次）", "线路强迫停运时间（小时）"):
                        ProjectComponentManager.add_project_component(wbs_code=wbs_code,
                                                                      name=component_name,
                                                                      params={name: [{"year": year, "value": value}]})

        return project_wbscode

    def aggregate_param(self, project_detail, aggr_filed):
        result = []
        need_delete_list = []
        for k, v in project_detail.items():
            if k.startswith(aggr_filed):
                need_delete_list.append(k)
                year = k[len(aggr_filed):].strip("/")
                if v is None:
                    continue
                result.append({"year": year, "value": v})

        for need_delete in need_delete_list:
            del project_detail[need_delete]

        project_detail[aggr_filed] = result

    def get_header(self, table):
        new_header = []
        header_list = []

        for i in range(1, table.max_row + 1):
            header = [item.value for item in table[i]]
            if header[0] and "项目名称" not in header[0] and i != 1:
                break
            header_list.append(header)
        header_row_count = len(header_list)

        for i in range(0, header_row_count):
            header = header_list[i]
            for j in range(0, len(header)):
                col_all_empty = True
                for k in range(i + 1, header_row_count):
                    if header_list[k][j] is not None:
                        col_all_empty = False

                if header[j] is None and not col_all_empty:
                    header[j] = header[j - 1]

        for items in zip(*header_list):
            new_header.append("/".join(str(item) for item in items).strip("/"))
        new_header = [item.replace("/None", "") for item in new_header]

        return new_header, header_row_count

    def parse_sheet(self, table):
        new_headers, header_row_count = self.get_header(table)

        for row in table[header_row_count + 1: table.max_row]:
            value = [item.value if item.value != -1 else None for item in row]
            project = dict(zip(new_headers, value))

            filtered_project = {}
            for name in project:
                if name \
                        and not name.startswith("公式列") \
                        and project[name] is not None:
                    filtered_project[name.strip()] = project[name]

            yield filtered_project


if __name__ == "__main__":
    # dataParser = DataParser()
    # with open(r"C:\Users\hangzhang2\Desktop\项目\后评价\项目后评价资料包\2021年数据\2021年江苏省电力公司投资效益数据-收资表2021.6 （-1提交系统）.xlsx",
    #           "rb") as old_project_file:
    #     wb = openpyxl.load_workbook(BytesIO(old_project_file.read()), data_only=True)
    #
    #     print(wb.sheetnames)

    with open(r"C:\Users\hangzhang2\Desktop\项目\后评价\项目后评价资料包\2021年数据\2021年江苏项目后评价数据0805最终上报.xlsx",
              "rb") as new_project_file:
        # dataParser.parse_old_project(BytesIO(old_project_file.read()))
        wb = openpyxl.load_workbook(BytesIO(new_project_file.read()), data_only=True)
